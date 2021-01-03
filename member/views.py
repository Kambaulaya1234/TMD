from django.http import HttpResponse, Http404, StreamingHttpResponse
from django.http import FileResponse
import os
from pathlib import Path
from django.conf import settings
from django.shortcuts import render, redirect
from .models import Member
from .forms import MemberForm
from django.views import View
from django.contrib import messages
from django.contrib.auth.models import User
from loan.models import Loan
from share.models import Share, WeekModel, YearModel
from accounts.models import Credential
import random
import string
from accounts.models import Profile
from openpyxl import load_workbook
import datetime
import pathlib


class IndexView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'members/member_list.html'
        members = Member.objects.filter(is_member=True, has_fine=False)
        form = MemberForm()
        context = {
            'members': members,
            'form': form,
        }
        return render(request, template_name, context)

    def post(self, *args, **kwargs):
        form = MemberForm(self.request.POST or None)
        if form.is_valid():
            object = form.save(commit=False)
            object.is_member = True
            object.save()
            messages.success(self.request, 'Your member created successfully!')
            return redirect('member:index')
        else:
            messages.error(self.request, 'Validation error')
            return redirect('member:index')


class InactiveIndexView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'members/inactive_member_list.html'
        members = Member.objects.filter(is_member=True, has_fine=True)
        form = MemberForm()
        context = {
            'members': members,
            'form': form,
        }
        return render(request, template_name, context)

from loan.models import LoanRound
class BaseObject:
    def get_object(self, id):
        member = Member.objects.filter(id=id)
        return member

    def get_all(self):
        members = Member.objects.all()
        loanRound = LoanRound.objects.all().delete()
        for member in members:
            member.user.delete()

    def members_list(self):
        members = Member.objects.filter(has_fine=True)
        return members


class StatusView(View, BaseObject):
    def post(self, request, *args, **kwargs):
        print(request.POST)
        week_ID = request.POST.get('week_id')
        year_ID = request.POST.get('year_id')
        member = Member.objects.filter(id=kwargs['id'])[0]
        _week = WeekModel.objects.filter(id=week_ID)[0]
        _year = YearModel.objects.filter(id=year_ID)[0]
        share = Share.objects.filter(member=member, week=_week, year=_year)[0]
        # ------------------------------------------------------
        member.has_fine = False
        member.fine_count = 0
        share.fine = 20000
        share.hisa = 1
        share.jamii = 5000
        # ------------------------------------------------------
        member.save()
        share.save()
        # ------------------------------------------------------
        messages.success(request, 'member activated successfully!')
        if not self.members_list().count():
            redirect_response = redirect('member:index')
        else:
            redirect_response = redirect('member:inactive')

        return redirect_response


class RemoveView(View, BaseObject):
    def get(self, *args, **kwargs):
        self.get_object(kwargs['id'])[0].user.delete()
        messages.success(self.request, 'member deleted successfully!')
        return redirect('member:index')


class EditMemberView(View, BaseObject):
    def post(self, *args, **kwargs):
        if self.get_object(kwargs['id']).exists():
            _member = self.get_object(kwargs['id'])
            context = {
                'username': self.request.POST['username'].lower(),
                'first_name': self.request.POST['first_name'],
                'last_name': self.request.POST['last_name'],
                'email': self.request.POST['email'],
            }
            if User.objects.filter(id=_member[0].user.id).exists():
                _phone = int(self.request.POST['phone'])
                user = User.objects.filter(id=_member[0].user.id)
                user.update(**context)
                Profile.objects.filter(
                    user=_member[0].user).update(phone=_phone)
                messages.success(self.request, 'member updated successfully!')
        return redirect('member:index')

from loan.models import LoanRound
class PayLoanView(View, BaseObject):
    def post(self, *args, **kwargs):
        member = self.get_object(kwargs['id'])[0]
        amount = float(self.request.POST['amount'])
        start = self.request.POST['start']
        deadline = self.request.POST['deadline']
        loanRound = self.request.POST['round']
        loanRoundOBJ,status=LoanRound.objects.get_or_create(name=loanRound)
        Loan.objects.create(member=member, amount=amount,
                            deadline=deadline, created_at=start,loan_round=loanRoundOBJ)
        messages.success(
            self.request, f'member {member.user} assigned loan  successfully!')
        return redirect('loan:index')


class CreateMemberView(View):
    @property
    def passcode(self):
        return ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))

    def post(self, *args, **kwargs):
        context = {
            'username': self.request.POST['last_name'].lower(),
            'first_name': self.request.POST['first_name'],
            'last_name': self.request.POST['last_name'],
            'email': self.request.POST['email'],
            'password': f'{self.passcode}'
        }
        if not User.objects.filter(username=context['username']).exists():
            _member_passcode = context['password']
            _phone = self.request.POST['phone']
            _user = User.objects.create_user(**context)
            Profile.objects.create(user=_user, phone=_phone, is_member=True)
            _member = Member.objects.create(user=_user, is_member=True)
            Credential.objects.get_or_create(
                member=_member, password=_member_passcode)

            messages.success(
                self.request, f'member {_user.get_full_name() } created   successfully!')
        else:
            messages.error(
                self.request, f"Member {context['username']} already exists!")
        return redirect('member:index')


class ImportMemberView(View):
    def post(self, *args, **kwargs):
        _member_file = self.request.FILES['file']
        if _member_file.name.endswith('.xlsx'):
            workbook = load_workbook(_member_file.file)
            sheets = workbook.sheetnames
            status = 0
            if 'members' in sheets:
                member_sheet = workbook.active
                if member_sheet.title == 'members':
                    for row in member_sheet.iter_rows(min_row=11, values_only=True):
                        print(*row)
                        _create = self.create_member_from_file(*row)
                else:
                    status = 1
                    messages.error(
                        self.request, f"Please set 'members sheet' as active in your excel ")
            if status == 0:
                messages.success(
                    self.request, f'members imported   successfully!')
        else:
            messages.error(
                self.request, f'Only excel (.xlsx) file is supported')
        return redirect('member:index')

    @property
    def passcode(self):
        return ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))

    def create_member_from_file(self, reg, firstname, sirname, lastname, email, phone):
        context = {
            'username': reg.lower(),
            'first_name': firstname,
            'last_name': lastname,
            'email': email,
            'password': f'{self.passcode}'
        }
        if not User.objects.filter(username=context['username']).exists():
            _phone = phone
            _member_passcode = context['password']
            _user = User.objects.create_user(**context)
            Profile.objects.create(user=_user, phone=_phone, is_member=True,
                                   sirname=sirname, ref_number=context['password'].upper())
            _member, status = Member.objects.get_or_create(
                user=_user, is_member=True)
            Credential.objects.get_or_create(
                member=_member, password=_member_passcode)

            # messages.success(
            #     self.request, f'member {_user} created   successfully!')
        else:
            messages.error(
                self.request, f"Member {context['username']} already exists!")
        return redirect('member:index')


class MemberProfileView(View, BaseObject):
    template_name = 'members/member_profile.html'

    def get(self, *args, **kwargs):
        MEMBER_ID = kwargs.get('id')
        _member = self.get_object(MEMBER_ID)[0]
        context = {
            'member': _member,
        }
        return render(self.request, self.template_name, context)
    
    
class MemberFinesView(View, BaseObject):
    template_name = 'members/member_fine.html'
    def get(self, *args, **kwargs):
        MEMBER_ID = kwargs.get('id')
        _member = self.get_object(MEMBER_ID)[0]
        context = {
            'member': _member,
        }
        return render(self.request, self.template_name, context)


class RemoveAll(View, BaseObject):
    def get(self, *args, **kwargs):
        self.get_all()
        messages.success(self.request, 'all member deleted successfully!')
        return redirect('member:index')


class Download(View):
    name_file = 'TEGEMEO DATABASE FORMATS.xlsx'

    def load(self, root):
        media = Path(settings.MEDIA_ROOT).resolve()
        file = media / 'FORMATS' / root
        _name = os.path.basename(file)
        if os.path.exists(file):
            with open(file, 'rb') as file_object:
                response = HttpResponse(file_object)
                response['Content-Type'] = 'application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response[
                    'Content-Disposition'] = f'attachment; filename={_name}'

                return response
        raise Http404

    def get(self, *args, **kwargs):
        return self.load(self.name_file)
