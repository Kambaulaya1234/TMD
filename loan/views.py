from django.shortcuts import render
from django.shortcuts import render, redirect
from django.views import View
from .models import Loan, Payment, LoanRound
from django.core.paginator import Paginator
from django.contrib import messages
from openpyxl import load_workbook
from member.models import Member
from django.db.models import Avg, Sum


class IndexView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'loans/loan_index.html'
        loans = Loan.objects.filter(status=False).order_by('-created_at')
        rounds = LoanRound.objects.order_by('-date')
        paginator = Paginator(rounds, 1)
        page = request.GET.get('page')
        filterItems = paginator.get_page(page)
        context = {
            'rounds': filterItems,
        }
        return render(request, template_name, context)


class CompleteLoanListView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'loans/loan_complete_index.html'
        rounds = LoanRound.objects.order_by('-date')
        paginator = Paginator(rounds, 1)
        page = request.GET.get('page')
        filterItems = paginator.get_page(page)
        context = {
            'rounds': filterItems,
        }
        return render(request, template_name, context)


class BaseObject:
    def get_object(self, id):
        loan = Loan.objects.filter(id=id)
        return loan

    def get_all(self):
        loans = Loan.objects.all().delete()

    def create_loan_payment(self, loan, paid, user, paid_date):
        _payment = Payment.objects.create(
            loan=loan, paid=paid, by=user, date=paid_date)
        if _payment.paid == 0:
            _payment.delete()
        return _payment

    def round_col(self, x):
        result = 5*x-2
        return result

    def net_payment(self, loan, paid_now, date_paid):
        _before = Payment.objects.filter(loan=loan).aggregate(total=Sum('paid'))
        if _before['total'] is None:
            _before['total'] = 0.0

        if _before['total'] < paid_now and paid_now > 0:
            _net = float(paid_now) - float(_before['total'])
            self.create_loan_payment(loan=loan, paid=float(_net),
                                     user=self.request.user, paid_date=date_paid)


class RemoveView(View, BaseObject):
    def get(self, *args, **kwargs):
        self.get_object(kwargs['id']).delete()
        messages.success(self.request, 'loan deleted successfully!')
        return redirect('loan:index')


class LoanProgressPaymentView(View, BaseObject):
    def post(self, *args, **kwargs):
        loan = self.get_object(kwargs['id'])[0]
        _paid = self.request.POST['paid']
        _paid_date = self.request.POST['paid_date']
        if loan.paid <= (loan.return_amount):
            if float(_paid) <= (loan.return_amount - loan.paid):
                loan.paid += float(_paid)
                self.create_loan_payment(
                    loan, _paid, self.request.user, _paid_date)
                loan.save()
                if loan.paid == loan.return_amount:
                    loan.status = True
                    loan.paid = loan.return_amount
                    loan.save()
                messages.success(
                    self.request, f'TZS {_paid}/= paid to {loan.member.user.get_full_name() } loan  successfully!')
            else:
                messages.warning(
                    self.request, f'Please consider only remained Amount (TZS { loan.return_amount - loan.paid}/=)')
        return redirect('loan:index')


class LoanShowView(View, BaseObject):
    def get(self, *args, **kwargs):
        _ID = kwargs['id']
        _loan = self.get_object(_ID)[0]
        template_name = 'loans/loan_show_payment.html'
        context = {
            'loan': _loan,
        }
        return render(self.request, template_name, context)


class RemovePaymentView(View, BaseObject):
    def get(self, *args, **kwargs):
        _ID = kwargs['id']
        _loan_id = self.request.GET['loan_id']
        loan = self.get_object(id=_loan_id)[0]
        payment_object = Payment.objects.filter(id=_ID)[0]
        if self.request.GET.get('discount'):
            _paid_amount = float(payment_object.paid)
            loan.paid -= _paid_amount
            loan.save()
            messages.success(
                self.request, f'TZS {payment_object.paid}/= discounted from {loan} loan successfully!')
        else:
            messages.success(
                self.request, f'record of TZS {payment_object.paid}/= from {loan} loan cleared successfully!')

        Payment.objects.filter(id=_ID).delete()
        return redirect('loan:show', loan.id)


class RemoveAll(View, BaseObject):
    def get(self, *args, **kwargs):
        _round=self.request.GET.get('round')
        if _round is None:
            LoanRound.objects.all().delete()
            messages.success(self.request, 'all Rounds  deleted successfully!')
        else:
            LoanRound.objects.filter(name=_round).delete()
            messages.success(self.request, f'all Round {_round}  deleted successfully!')
        return redirect('loan:index')


class HandleUploadedFileView(View, BaseObject):
    def post(self, *args, **kwargs):
        _loan_file = self.request.FILES['file']
        if _loan_file.name.endswith('.xlsx'):
            workbook = load_workbook(_loan_file.file)
            sheet = workbook.active
            sheets = workbook.sheetnames
            status_sms = 0
            if 'loans' in sheets:
                loan_sheet = workbook.active
                if loan_sheet.title == 'loans':
                    # =========================main operation=====================================

                    current_year = loan_sheet["B4"].value
                    current_month = loan_sheet["B5"].value
                    ending_round = loan_sheet["D6"].value

                    if ending_round is None or int(ending_round) == 0:
                        ending_round = loan_sheet["B6"].value
                    all_ranges = [x for x in range(
                        int(loan_sheet["B6"].value), int(ending_round)+1)]

                    # =========================Main operation=====================================
                    for code in all_ranges:
                        active_round_code = self.round_col(code)+4
                        round_name = loan_sheet.cell(
                            row=9, column=self.round_col(code)+1).value
                        round_col_id = self.round_col(code)+1

                        for row in sheet.iter_rows(min_row=11, min_col=1, max_col=active_round_code, values_only=True):
                            username, *_, _amount, _start_date, _deadline, _paid, p_date = row
                            FILTERED_ROW = username, _amount, _start_date, _deadline, _paid, round_name, p_date
                            if (_amount is not None and _amount > 0):
                                _created, _status = self.create_loan_from_file(
                                    *FILTERED_ROW)
                                if not _status:
                                    status_sms = 1
                                    messages.error(
                                        self.request, f'member "{_created}" is not found in member list!')

                    # =========================End operation=====================================

                else:
                    status_sms = 1
                    messages.error(
                        self.request, f"Please set 'loans sheet' as active in your excel ")

            else:
                status_sms = 1
                messages.error(
                    self.request, f"No 'loans sheet' inside your excel workbook")
        else:
            status_sms = 1
            messages.error(
                self.request, f'Only excel (.xlsx) file is supported')

        if status_sms == 0:
            messages.success(
                self.request, f'loans imported successfully')
        return redirect('loan:index')

    def create_loan_from_file(self, member_entry,
                              amount_entry, start_entry,
                              deadline_entry, paid_entry,
                              round_entry, p_date):
        _member = Member.objects.filter(
            user__username__icontains=member_entry.lower()).distinct()
        if _member.exists() and amount_entry > 0:
            roundObject, info = LoanRound.objects.update_or_create(
                name=round_entry)
            context = {
                'member': _member[0], 'amount': float(amount_entry), 'created_at': start_entry,
                'deadline': deadline_entry, 'loan_round': roundObject
            }

            _loan, info = Loan.objects.update_or_create(**context)
            if paid_entry is not None and paid_entry > 0:
                self.net_payment(
                    loan=_loan, paid_now=paid_entry, date_paid=p_date)
                _loan.paid = float(paid_entry)
                _loan.save()

            return (member_entry, True)
        else:
            return (member_entry, False)
