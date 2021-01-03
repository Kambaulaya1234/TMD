from django.shortcuts import render, redirect
from django.views import View
from .models import Share, WeekModel, MonthModel, YearModel,Fine
from .forms import CreateShareForm
from django.contrib import messages
from django import forms
from django.core.paginator import Paginator
from datetime import datetime
from member.models import Member
from openpyxl import load_workbook
from django.core.files.base import ContentFile


class IndexView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'share/share_index.html'
        if (MonthModel.objects.exists() and YearModel.objects.exists()):
            this_month = MonthModel.objects.order_by('-created_at')[0]
            this_year = YearModel.objects.order_by('-created_at')[0]
            weeks = WeekModel.objects.filter(
                month__lte=this_month, year__lte=this_year).order_by('-created_at')
            paginator = Paginator(weeks, 1)
            page = request.GET.get('page')
            filterItems = paginator.get_page(page)
            _initial = {
                'jamii': 5000
            }
            form = CreateShareForm(initial=_initial)
            context = {
                'weeks': filterItems,
                'form': form,
            }
            return render(request, template_name, context)
        else:
            return render(request, template_name)

    def post(self, *args, **kwargs):
        form = CreateShareForm(self.request.POST or None)
        if form.is_valid():
            object = form.save(commit=False)
            if(object.hisa < 1 or object.jamii < 5000):
                object.member.has_fine = True
                object.member.save()
            object.save()
            messages.success(self.request, 'Your share created successfully!')
            return redirect('share:index')
        else:
            messages.error(self.request, '')
            return redirect('share:index')


class BaseObject:
    def get_object(self, id):
        share = Share.objects.filter(id=id)
        return share

    def get_all(self):
        share = Share.objects.all()
        return share


class Update(View, BaseObject):
    def post(self, *args, **kwargs):
        share = self.get_object(kwargs['slug'])[0]
        form = CreateShareForm(self.request.POST or None, instance=share)
        if form.is_valid():
            form.save()
            messages.success(self.request, 'Your post updated successfully!')
            return redirect('share:index')
        else:
            messages.warning(self.request, 'Error Validation successfully!')
            return redirect('share:index')


class RemoveView(View, BaseObject):
    def get(self, *args, **kwargs):
        self.get_object(kwargs['id']).delete()
        messages.success(self.request, 'Share deleted successfully!')
        return redirect('share:index')


class RemoveAll(View, BaseObject):
    def get(self, *args, **kwargs):
        self.get_all().delete()
        WeekModel.objects.all().delete()
        MonthModel.objects.all().delete()
        YearModel.objects.all().delete()
        messages.success(self.request, 'all Share deleted successfully!')
        return redirect('share:index')


class ActiveWeekView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'share/share_week_index.html'
        weeks = WeekModel.objects.order_by('-created_at')
        paginator = Paginator(weeks, 10)
        page = request.GET.get('page')
        filterItems = paginator.get_page(page)
        form = CreateShareForm()
        context = {
            'weeks': filterItems,
            'form': form,
        }
        return render(request, template_name, context)


class ActiveMonthView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'share/share_month_index.html'
        weeks = MonthModel.objects.order_by('-created_at')
        paginator = Paginator(weeks, 10)
        page = request.GET.get('page')
        filterItems = paginator.get_page(page)
        form = CreateShareForm()
        context = {
            'months': filterItems,
            'form': form,
        }
        return render(request, template_name, context)


class ActiveYearView(View):
    def get(self, request, *args, **kwargs):
        template_name = 'share/share_year_index.html'
        weeks = YearModel.objects.order_by('-created_at')
        paginator = Paginator(weeks, 10)
        page = request.GET.get('page')
        filterItems = paginator.get_page(page)
        form = CreateShareForm()
        context = {
            'years': filterItems,
            'form': form,
        }
        return render(request, template_name, context)


class HandleUploadedFileView(View):
    def week_col(self, index):
        result = 2*index + 2
        return result

    def month_name_detector(self, sheet_name, code_name):
        month_name = sheet_name.cell(
            row=9, column=self.week_col(code_name)-1).value
        if '.' in month_name:
            return (month_name.split('.')[1], True)
        elif '/' in month_name:
            return (month_name.split('/')[1], True)
        else:
            return (month_name, False)

    def post(self, *args, **kwargs):
        # form = UploadFileForm(self.request.POST or None,
        #                       self.request.FILES or None)
        # if form.is_valid():
        #     form.save()
        #     messages.error(
        #         self.request, f'Error with file format,Please wait')
        _share_file = self.request.FILES['file']
        if _share_file.name.endswith('.xlsx'):
            workbook = load_workbook(_share_file.file)
            sheet = workbook.active
            sheets = workbook.sheetnames
            status_sms = 0
            if 'shares' in sheets:
                share_sheet = workbook.active
                if share_sheet.title == 'shares':
                    # =========================main operation=====================================
                    current_year = share_sheet["B4"].value

                    ending_week = share_sheet["D6"].value

                    if ending_week is None or int(ending_week) == 0:
                        ending_week = share_sheet["B6"].value
                    all_ranges = [x for x in range(
                        int(share_sheet["B6"].value), int(ending_week)+1)]
                    # =========================Main operation=====================================
                    for code in all_ranges:
                        active_week_code = self.week_col(code)
                        week_name = share_sheet.cell(
                            row=8, column=self.week_col(code)-1).value
                        current_month, status = self.month_name_detector(
                            share_sheet, code)
                        if not status:
                            status_sms = 1
                            messages.error(
                                self.request, f'Please consider fullstop "." or Forward slash "/" in {current_month} date separation ')
                        else:
                            for row in sheet.iter_rows(min_row=11, min_col=1, max_col=active_week_code, values_only=True):
                                username, *_, _hisa, _jamii = row
                                FILTERED_ROW = username, _hisa, _jamii, current_year, current_month, week_name
                                _created, _status = self.create_member_from_file(
                                    *FILTERED_ROW)
                                if not _status:
                                    status_sms = 1
                                    messages.error(
                                        self.request, f'member "{_created}" is not found in member list!')

                    # =========================End operation=====================================

                else:
                    status_sms = 1
                    messages.error(
                        self.request, f"Please set 'shares sheet' as active in your excel ")

            else:
                status_sms = 1
                messages.error(
                    self.request, f"No 'shares sheet' inside your excel workbook")
        else:
            status_sms = 1
            messages.error(
                self.request, f'Only excel (.xlsx) file is supported')

        if status_sms == 0:
            messages.success(
                self.request, f'Share imported successfully')
        return redirect('share:index')

    def create_member_from_file(self, member_entry, hisa, jamii, year_entry, month_entry, week_entry):
        _member = Member.objects.filter(
            user__username__icontains=member_entry.lower()).distinct()
        if _member.exists():
            _year, info = YearModel.objects.get_or_create(name=year_entry)
            _month, info = MonthModel.objects.get_or_create(
                name=month_entry, year=_year)
            _week, info = WeekModel.objects.get_or_create(
                name=week_entry, month=_month, year=_year)

            context = {
                'member': _member[0], 'week': _week, 'month': _month,
                'year': _year, 'hisa': hisa, 'jamii': jamii
            }

            _share, info = Share.objects.get_or_create(**context)
            if(_share.hisa < 1 or _share.jamii < 5000):
                if _share.member.fine_count >= 3:
                    _share.member.has_fine = True
                    _share.member.save()
                else:
                    _share.member.fine_count += 1
                    if _share.hisa < 0 and _share.jamii < 5000:
                        Fine.objects.create(member=_share.member,hisa_amount=10000,jamii_amount=10000,week=_share.week)
                    elif  _share.hisa < 0:
                        Fine.objects.create(member=_share.member,hisa_amount=10000,jamii_amount=0,week=_share.week)
                        
                    elif  _share.jamii < 5000:
                        Fine.objects.create(member=_share.member,hisa_amount=0,jamii_amount=10000,week=_share.week)
                        
                    _share.member.save()

            return (member_entry, True)
        else:
            return (member_entry, False)
